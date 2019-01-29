'''
阴阳师式神传记爬虫
'''

import re
from selenium import webdriver
from openpyxl import Workbook
from openpyxl import load_workbook

class Spider():
    # 指定网页地址
    url = 'http://yys.163.com/shishen/'

    # 指定导出的Excel表的名称
    xlsx_name = 'yys_biography.xlsx'

    # 正则表达式
    root_filter = '<div class="shishen_wrap clearfix">([\s\S]*?)</a></div></div>'
    list_filter = '<a href="([\s\S]*?)">'

    name_filter = '<input type="hidden" id="shishenName" value="([\s\S]*?)" />\n    <input type="hidden"'
    
    data_lv_filter = '<div class="person_name">([\s\S]*?)</span></div>'
    lv_filter = '<span class="level ([\s\S]*?)">'

    # data_biography_filter = '<div class="zhuanji_txt_wrap([\s\S]*?)</div>'
    data_biography_filter2 = '<div id="section1">([\s\S]*?)</div>\n            </div>'
    biography_filter = 'px;">([\s\S]*?)<p class="japanese"><span>'
    
    # 获取原始数据页
    def __fetch_list(self):
        # 声明浏览器
        option_chrome = webdriver.ChromeOptions()
        option_chrome.add_argument('--no-sandbox')                          # 给予root权限
        option_chrome.add_argument('--headless')                            # 不显示页面
        option_chrome.add_argument('--disable-gpu')                         # 谷歌文档提到需要加上这个属性来规避bug
        option_chrome.add_argument('blink-settings=imagesEnabled=false')    # 不加载图片, 提升速度
        browser = webdriver.Chrome(chrome_options = option_chrome)

        # browser = webdriver.Chrome()    # 声明浏览器
        browser.get(Spider.url)         # 访问网页
        htmls = browser.page_source     # 将源码赋值给htmls
        browser.close()                 # 关闭浏览器

        return htmls

    # 从原始数据中，截取式神名称与对应的编号并返回
    def __analysis_list(self, htmls):
        root_html = re.findall(Spider.root_filter, htmls)

        shikigami_list = []

        for htmls in root_html:
            shikigami_list = re.findall(Spider.list_filter, htmls)

        return shikigami_list

    # 获取式神数据
    def __fetch_data(self, shikigami_list):
        # 声明浏览器
        option_chrome = webdriver.ChromeOptions()
        option_chrome.add_argument('--no-sandbox')                          # 给予root权限
        option_chrome.add_argument('--headless')                            # 不显示页面
        option_chrome.add_argument('--disable-gpu')                         # 谷歌文档提到需要加上这个属性来规避bug
        option_chrome.add_argument('blink-settings=imagesEnabled=false')    # 不加载图片, 提升速度
        browser = webdriver.Chrome(chrome_options = option_chrome)

        # browser = webdriver.Chrome()               # 声明浏览器
        browser.get(Spider.url + shikigami_list)     # 访问网页
        data = browser.page_source                   # 将源码赋值给data
        browser.close()                              # 关闭浏览器

        # 截取式神数据中的名称、品质、传记，并输出到Excel

        name = re.findall(Spider.name_filter, data)
        name = str(name[0])
        name = re.sub('&amp;', '&', name)

        data_lv = re.findall(Spider.data_lv_filter, data)
        data_lv = str(data_lv[0])
        lv = re.findall(Spider.lv_filter, data_lv)
        lv = str(lv[0])

        data_biography = re.findall(Spider.data_biography_filter2, data)
        if data_biography == ['\n            <div class="section_container">\n                <h3 class="com_title title1"></h3>\n                <div class="zhuanji_container clearfix"><p class="no_story">该式神无传记</p>']:
            biography = ['该式神无传记']
        else:
            data_biography = str(data_biography[0])
            biography = re.findall(Spider.biography_filter, data_biography)
        biography = str(biography[0])
        biography = re.sub(' ', '', biography)
        biography = re.sub('<pclass="chineseshow"><span>', '', biography)
        biography = re.sub('<br/></span></p>', '', biography)
        biography = re.sub('﹁', '"', biography)
        biography = re.sub('﹂', '"', biography)

        shikigami_info = (name, lv, biography)
        self.__append_xlsx(Spider.xlsx_name, shikigami_info)

    # 创建Excel
    def __active_xlsx(self, xlsx_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "式神信息"
        top_bar = ('式神名称', '式神品质', '式神介绍')
        ws.append(top_bar)
        wb.save(xlsx_name)

    # 写入数据
    def __append_xlsx(self, xlsx_name, data):
        wb = load_workbook(xlsx_name)
        sheet = wb["式神信息"]
        sheet.append(data)
        wb.save(xlsx_name)


    # 调用私有方法的入口方法
    # 同时也是执行流程控制
    def go(self):
        self.__active_xlsx(Spider.xlsx_name)                            # 创建Ecxel表
        htmls = self.__fetch_list()                                     # 获取原始网页
        shikigami_list = self.__analysis_list(htmls)                    # 从原始网页中提取式神列表
        shikigami_data = map(self.__fetch_data, shikigami_list)         # 遍历式神列表的每一项，提取信息，输出到Excel
        list(shikigami_data)
        # self.__fetch_data('262.html')

spider = Spider()
spider.go()
